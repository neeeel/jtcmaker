from tkinter import filedialog
from PIL import Image,ImageDraw,ImageTk,ImageFont
import mapmanager
import pandas as pd
from tkinter import filedialog,messagebox
import openpyxl
import os
import xlwt
from openpyxl.chart import LineChart,PieChart,BarChart
import math
import datetime
import pickle
import copy

outline = 3 ## witdth of line to draw circles on the map
siteDetails = None
baseClasses = [["Car",1],["LGV",1],["OGV1",1.5],["OGV2",2.3],["PSV",2],["MC",0.4],["PC",0.2]]
projectClasses = [["Car",1],["LGV",1],["OGV1",1.5],["OGV2",2.3],["PSV",2],["MC",0.4],["PC",0.2]]
excelMapWidth = 576.5295 ### the size of the map in the excel template sheet
excelMapHeight = 576.5295
sites = {}
groups = {}
groupCount = 1
current_label = 64
individual_site_zoom_value = 19
overview_map_details = []
jobNumber = ""
jobName = ""
surveyDate= ""
timePeriods  = ""
project = None



####################################################################################################
###
### Deal with classes
###
####################################################################################################

def get_classes(surveyIndex):
    surveyType = ["J","Q","P"][surveyIndex]
    return project["surveys"][surveyType]["classes"]

def delete_class(index,surveyIndex):
    surveyType = ["J", "Q", "P"][surveyIndex]
    project["surveys"][surveyType]["classes"].pop(index)
    return len(project["surveys"][surveyType]["classes"])

def add_class(vals,surveyIndex):
    surveyType = ["J", "Q", "P"][surveyIndex]
    project["surveys"][surveyType]["classes"].append(vals)
    return len(project["surveys"][surveyType]["classes"])

def edit_class(vals,index,surveyIndex):
    surveyType = ["J", "Q", "P"][surveyIndex]
    project["surveys"][surveyType]["classes"][index] = vals
    return len(project["surveys"][surveyType]["classes"])

def move_class_up(index,surveyIndex):
    surveyType = ["J", "Q", "P"][surveyIndex]
    if index == 0:
        return index
    temp = project["surveys"][surveyType]["classes"][index]
    project["surveys"][surveyType]["classes"][index] = project["surveys"][surveyType]["classes"][index-1]
    project["surveys"][surveyType]["classes"][index - 1] = temp
    return index -1

def move_class_down(index,surveyIndex):
    surveyType = ["J", "Q", "P"][surveyIndex]
    if index == len(project["surveys"][surveyType]["classes"])-1:
        return index
    temp = project["surveys"][surveyType]["classes"][index]
    project["surveys"][surveyType]["classes"][index] = project["surveys"][surveyType]["classes"][index+1]
    project["surveys"][surveyType]["classes"][index + 1] = temp
    return index + 1

####################################################################################################
###
### Deal with groups
###
####################################################################################################

def add_group():
    global groupCount, groups
    groupName = "Group " + str(groupCount)
    groupCount+=1
    groups[groupName] = {}
    groups[groupName]["siteList"] = []
    groups[groupName]["coords"] = []
    return groupName

def delete_group(groupName):
    global groups,groupCount
    del groups[groupName]
    groupCount-=1
    print(sorted(groups.items(), key=lambda x:0 if x[0] == "ALL" else int(x[0].replace("Group ","").strip())))
    for index,key in enumerate(sorted(groups.keys(), key=lambda k:0 if k == "ALL" else int(k.replace("Group ","").strip()))):
        print(index,key)
        if key != "ALL":
            groups["Group " + str(index)] = groups.pop(key)

def add_site_to_group(groupName,site):
    global groups
    groups[groupName]["siteList"].append(site)

def delete_site_from_group(groupName,site):
    global groups
    groups[groupName]["siteList"].remove(site)

def get_groups():
    return groups


####################################################################################################
###
###  Deal with Project
###
####################################################################################################


####################################################################################################
###
###  Deal with sites
###
####################################################################################################


def get_site_by_order(index):
    for _,site in project["sites"].items():
        if site["order"] == index:
            return site
    return None



def change_site_image_type(val,site):
    if val ==0:
        val  = "roadmap"
    else:
        val = "satellite"
    if site["imageType"] != val:
        x,y = site["latlon"]
        map = mapmanager.load_high_def_map_with_labels(x, y, site["zoom"],imageType=val)
        map.save(str(site["Site Name"]) + ".png")
        site["imageType"] = val
        site["image"] = map.resize((800, 800), Image.ANTIALIAS)


def change_site_zoom(site,value):
    if value == "+":
        site["zoom"] += 1
    else:
        site["zoom"]-=1
    map = mapmanager.load_high_def_map_with_labels(site["latlon"][0], site["latlon"][1], site["zoom"],imageType=site["imageType"])
    map.save(str(site["Site Name"]) + ".png")
    site["image"] = map.resize((800, 800), Image.ANTIALIAS)
    armlist = [k for k, item in site["Arms"].items()]
    for arm in armlist:
        delete_arm_from_site(site["Site Name"], arm)

def change_site_centre_point(site,x,y):
    ###
    ### x,y are the deltas that the map has changed by
    ###
    print("changing site centre point,movement was",x,y)
    currentCentre = sites[site]["latlon"]
    zoom = sites[site]["zoom"]
    newCentre = mapmanager.get_lat_lon_from_x_y(currentCentre,(640-(x*800/1280)),(640-(y*800/1280)),zoom,size=1280)
    print("new centre is ",newCentre)
    site["latlon"] = newCentre
    download_overview_map()
    map = mapmanager.load_high_def_map_with_labels(sites[site]["latlon"][0], sites[site]["latlon"][1], sites[site]["zoom"],imageType =sites[site]["imageType"])
    map.save(str(sites[site]["Site Name"]) + ".png")
    site["image"] = map.resize((800, 800), Image.ANTIALIAS)
    armlist = [k for k, item in sites[site]["Arms"].items()]
    for arm in armlist:
        delete_arm_from_site(site["Site Name"],arm)

def change_site_group(site,group):
    sites[site]["group"] = group

def add_arm_to_site(siteName,x,y):
    global current_label,siteDetails,sites
    if siteDetails is None:
        return
    current_label += 1
    armName = chr(current_label)
    edit_arm(siteName, armName, x, y)
    return chr(current_label)

def delete_arm_from_site(siteName,armName):
    site = sites.get(siteName, {})
    print("before delete, site is",site)
    try:
        del site["Arms"][armName]
    except Exception as e:
        pass
        ### something went wrong, and the arm didnt exist
    print("after delete, site is", site)



####################################################################################################
###
###  Deal with Arms
###
####################################################################################################



def decrement_arm_label():
    global current_label
    ###
    ### a label was deleted in the user interface, need to decrement the counter so we give the correct letter
    ### when requested
    current_label-=1

def edit_arm(siteName,armName,x,y):
    print("in edit arm",siteName,armName,x,y)
    global siteDetails, sites
    site = sites.get(siteName, {})
    if site == {}:
        site["zoom"] = individual_site_zoom_value
    print("site is", site)
    site["coords"] = []
    row = siteDetails[siteDetails["Site Name"] == siteName]
    siteLatLon = row.values.tolist()[0][1:]
    print("sitelatlon is",siteLatLon)
    armLatLon = mapmanager.get_lat_lon_from_x_y(siteLatLon, x, y, site["zoom"])
    site["coords"]=mapmanager.get_coords(overview_map_details[1],siteLatLon,overview_map_details[2],size=1280)
    print("location for arm is", armLatLon)
    site["Arms"][armName] = {}
    site["Arms"][armName]["latlon"] = armLatLon
    site["Arms"][armName]["coords"] = (x, y)
    site["Arms"][armName]["road"] = mapmanager.get_road_name(armLatLon[0], armLatLon[1])
    print("road for arm is",site["Arms"][armName]["road"])
    site["Arms"][armName]["orientation"] = 0
    print("site is now", site)
    print("")
    sites[siteName] = site

def edit_arm_orientation(siteName,armName,orientation):
    global siteDetails, sites
    site = sites.get(siteName, {})
    #print("in edit arm, params are", siteName, armName, x, y)
    site["Arms"][armName]["orientation"] = orientation
    print("site is now", site)

def get_site_map(siteName):
    print("looking for ",str(siteName) + ".png")
    try:
        img = Image.open(str(siteName) + ".png")
        img = img.resize((800, 800), Image.ANTIALIAS)
        print("returning")
        return ImageTk.PhotoImage(img)
    except Exception as e:
        return None

def get_overview_map():
    try:
        img = Image.open("overview.png")
        #img = img.resize((800, 800), Image.ANTIALIAS)
        coordsList = [(site["Site Name"],mapmanager.get_coords(overview_map_details[1],site["latlon"],overview_map_details[2],size=1280)) for _,site in sites.items()]
        print("Coord list is",coordsList)
        return [img,coordsList]
    except Exception as e:
        print("error",e)
        return [None,[]]

def get_nearest_site_on_overview_map(x,y):
    pass

def get_project_details():
    return [project["jobName"],project["jobNumber"]]

def format_date(d):
    surveyDate = None
    if not type(d)== datetime.datetime:
        try:
            surveyDate = datetime.datetime.strptime(d,"%d/%m/%Y")
        except Exception as e:
            try:
                surveyDate = datetime.datetime.strptime(d, "%d/%m/%y")
            except Exception as e:
                pass
        return surveyDate
    else:
        return d


def import_site_details_from_excel():
    project = None
    global siteDetails,jobName,jobNumber,surveyDate,timePeriods,projectClasses, project
    f = filedialog.askopenfilename(initialdir=dir)
    if f == "":
        return
    siteDetails = pd.read_excel(f,parse_cols=[0,1,2,3],index_col=None)
    siteDetails["Site Name"] = siteDetails["Site Name"].apply(str)
    siteDetails["Site Name"] = siteDetails["Site Name"].apply(str.title)
    mask = ~siteDetails["Site Name"].str.contains("Site")
    siteDetails.loc[mask, "Site Name"] = siteDetails[mask]["Site Name"].apply(lambda x: "Site " + x)
    mask = ~siteDetails["Site Name"].str.contains("Site ")
    siteDetails.loc[mask, "Site Name"] = siteDetails[mask]["Site Name"].apply(lambda x: "Site " + x.split("Site")[1])
    siteDetails.dropna(axis=0, inplace=True)
    siteDetails[siteDetails["Type"].isnull()] = "J"
    print(siteDetails)

    wb = openpyxl.load_workbook(f)
    project = {}
    project["jobNumber"] = wb.worksheets[0]["F4"].value
    project["jobName"] = wb.worksheets[0]["F6"].value
    project["surveys"] = {}
    surveyTypes = ["J","P","Q"]
    for index,survey in enumerate(surveyTypes):
        project["surveys"][survey] = {}
        project["surveys"][survey]["date"] = format_date(wb.worksheets[0].cell(row=8+index, column=6).value)
        col = 6
        project["surveys"][survey]["classes"] = []
        while not wb.worksheets[0].cell(row=12+(2*index), column=col).value is None:
            try:
                pcu = int(wb.worksheets[0].cell(row=13 + (2*index),column=col).value)
            except Exception as e:
                pcu = 1
                project["surveys"][survey]["classes"].append([wb.worksheets[0].cell(row=12+(2*index),column=col).value,pcu])
            col+=1

        project["surveys"][survey]["times"] = ""
        col = 6
        timeList = []
        while not wb.worksheets[0].cell(row=19+index,column=col).value is None:
            print(wb.worksheets[0].cell(row=19+index,column=col).value)
            try:
                t = wb.worksheets[0].cell(row=19+index,column=col).value
                timeList.append(t.strftime("%H:%M"))
                print("timelist is",timeList)
            except Exception as e:
                messagebox.showinfo(message="there was a problem with one or more times in the project file\n Please check")
                project = None
                return
            col+=1
        if len(timeList) != 0:
            if len(timeList)%2!=0:
                timeList = timeList[:-1]
            project["surveys"][survey]["times"] = ",".join([str(timeList[i]) + "-" + str(timeList[i+1]) for i in range(0,len(timeList),2)])
    ###
    ### fill in any blank survey dates and times
    ###
    #t = [s["times"] for _,s in project["surveys"].items() if not s["times"] is ""]
    #d = [s["date"] for _,s in project["surveys"].items() if not s["date"] is None]
    #print(t,d)
    sites = {}
    project["sites"] = {}
    for index, site in siteDetails.iterrows():
        surveyTypes = site[3].split("/")
        project["sites"][site[0]] = {}
        project["sites"][site[0]]["order"] = index
        project["sites"][site[0]]["site Name"] = site[0]
        project["sites"][site[0]]["surveys"] ={}
        for survey in surveyTypes:
            project["sites"][site[0]]["surveys"][survey] = {}
            project["sites"][site[0]]["surveys"][survey]["Arms"] = {}
            project["sites"][site[0]]["surveys"][survey]["zoom"] = individual_site_zoom_value
            if survey.upper() == "Q":
                project["sites"][site[0]]["surveys"][survey]["imageType"] = "satellite"
            else:
                project["sites"][site[0]]["surveys"][survey]["imageType"] = "roadmap"
            project["sites"][site[0]]["surveys"][survey]["image"] = None
            project["sites"][site[0]]["surveys"][survey]["latlon"] = (site[1], site[2])


    print("project is",project)



def load_project():
    global siteDetails, groups,sites,projectClasses,project
    import_site_details_from_excel()
    if siteDetails is None:
        return None

    download_overview_map()
    for _,site in sites.items():
        site["coords"] = mapmanager.get_coords(overview_map_details[1],site["latlon"],overview_map_details[2],size=1280)
    download_all_individual_site_maps()
    groups["ALL"] = {}
    groups["ALL"]["siteList"] = [site["Site Name"] for _,site in sites.items()]
    groups["ALL"]["coords"] = []
    with open("test" + ".pkl", "wb") as f:
        pickle.dump(project, f)
    with open("test" + ".pkl", "rb") as f:
        project = pickle.load(f)
    print("after loading, project is",project)
    print(get_site_by_order(0))
    return get_site_by_order(0)




def save_project_to_pickle(file):
    with open(file + ".pkl","wb") as f:
        pickle.dump(project,f)
    global sites,groups,project
    file = file.replace(".pkl","")
    sitesToSave = {}
    for siteName,site in project["sites"].items():
        print(siteName,site)
        newSite  = {}
        sitesToSave[siteName] = newSite
        for key,value in site.items():
            print(key,value)
            if key == "Arms":
                newSite["Arms"] = {}
                for armLabel,arm in value.items():
                    print(armLabel,arm)
                    newSite["Arms"][armLabel] = {}
                    for k,v in arm.items():
                        if k == "label":
                            pass
                        else:
                            newSite["Arms"][armLabel][k] = v

                pass
            else:
                newSite[key] = value
    print("sites to save is",sitesToSave)
    print("sites are",sites)
    projectToSave = {"sites":sitesToSave,"groups":groups,"details":[project["jobName"],project["jobNumber"]],"classes":projectClasses}
    with open(file + ".pkl","wb") as f:
        pickle.dump(project,f)


def load_project_from_pickle(file):
    global sites,groups,jobName, jobNumber, surveyDate, timePeriods,groupCount,projectClasses,project
    with open("test" + ".pkl", "rb") as f:
        project = pickle.load(f)
    return get_site_by_order(0)



    if ".pkl" in file:
        try:
            with open(file, "rb") as f:
                project = pickle.load(f)
                print("loaded project is",project)
                sites = project["sites"]
                groups=project["groups"]
                groupCount = len(groups)
                details = project["details"]
                projectClasses = project["classes"]
                jobName, jobNumber, surveyDate, timePeriods = details
                download_overview_map()
                return get_site_by_order(0)
        except Exception as e:
            print(e)
            return None

    return None

def load_previous_site(site):
    ###
    ### user has pressed the left arrow to move to the previous site
    ###

    index = site["order"] - 1
    result = get_site_by_order(index)
    if result is None:
        return site
    return result

def load_next_site(site):
    ###
    ### user has pressed the right arrow to move to the next site
    ###
    index = site["order"] + 1
    result = get_site_by_order(index)
    if result is None:
        return site
    return result

def export_to_excel(jobDetails):
    global groups
    file = filedialog.asksaveasfilename()
    if file == "" or file is None:
        return
    wb = openpyxl.load_workbook("Flow Automation Template v2.0.xlsm",keep_vba=True)

    sht = wb.get_sheet_by_name("Maps")

    img = Image.open("Tracsis logo.png")
    imgSmall = img #.resize((247, 78), Image.ANTIALIAS)
    excelImageSmall = openpyxl.drawing.image.Image(imgSmall)
    sht.add_image(excelImageSmall, "AA1")

    img = Image.open("Tracsis Icon.png")
    imgSmall = img #.resize((25, 25), Image.ANTIALIAS)
    excelImageSmall = openpyxl.drawing.image.Image(imgSmall)
    sht.add_image(excelImageSmall, "AA2")

    img = Image.open("Tracsis-banner.png")
    imgSmall = img  # .resize((25, 25), Image.ANTIALIAS)
    excelImageSmall = openpyxl.drawing.image.Image(imgSmall)
    sht.add_image(excelImageSmall, "AA3")

    sht = wb.get_sheet_by_name("Dashboard")

    c1= BarChart()
    c1.title = "Class Ratios"
    sht.add_chart(c1,"AB5")

    c1 = BarChart()
    c1.title = "Total Volumes"
    sht.add_chart(c1, "M25")



    sht = wb.get_sheet_by_name("Data")
    sht.cell(row=2,column=14).value = ",".join(jobDetails["times"])
    sht.cell(row=3, column=14).value = jobDetails["period"]
    sht.cell(row=4, column=14).value = jobNumber
    sht.cell(row=5, column=14).value = jobName
    sht.cell(row=6,column=14).value = surveyDate.strftime("%d/%m/%Y")
    for index,cl in enumerate(projectClasses):
        sht.cell(row=index+2, column=50).value = cl[0]
        sht.cell(row=index + 2, column=52).value = cl[1]
    row = 2

    fnt = ImageFont.truetype("arial", size=18)

    count = 0
    for key,site in sorted(sites.items(),key=lambda x:x[1]["order"]): ### sorted(sites.items(),key=lambda x:int(x[0].replace("Site ","").strip())):

        ###
        ### check and adjust road names if any are duplicated within a site
        ###
        print("here")
        dets = [[k, i["orientation"] + 180, i["road"]] for k, i in site["Arms"].items()]
        mylist = [item["road"] for key, item in site["Arms"].items()]
        print([(item["road"],key) for key, item in site["Arms"].items()])
        print("list of roads",mylist)
        for road in mylist:
            matches = [x for x in dets if x[2] == road]
            if len(matches) > 1:
                print("matches are",matches)
                for m in matches:

                    road = m[2] + "(" + road_orientation(m[1]) + ")"
                    site["Arms"][m[0]]["road"] = road

        sht = wb.get_sheet_by_name("Maps")
        sht.cell(row=count+2,column=1).value = key
        sht.cell(row=count+2,column=13).value = "No relevant observations"
        sht = wb.get_sheet_by_name("Data")
        siteImg = Image.open(key + ".png").convert('RGB')
        siteImg = siteImg.resize((800, 800), Image.ANTIALIAS)
        drawimage = ImageDraw.Draw(siteImg)
        if len(site["Arms"]) !=0:
            col = 4
            sht.cell(row=row,column=1).value = key
            sht.cell(row=row , column=2).value = str(site["coords"][0]) + "," + str(site["coords"][1])
            sht.cell(row=row , column=3).value = str(site["latlon"][0]) + "," + str(site["latlon"][1])
            for label,arm in sorted(site["Arms"].items()):
                angle = arm["orientation"]
                x,y = arm["coords"]
                outline = 3  # line thickness

                drawimage.ellipse([x - 15- outline, y - 15- outline, x + 15+ outline, y + 15+ outline], outline="Black",fill = "black")
                drawimage.ellipse([x - 15, y - 15, x + 15, y + 15], outline="white",fill = "white")
                drawimage.text((x-6, y-7), text=label,font=fnt, fill="black")
                angle += 180
                if angle >= 360:
                    angle-=360
                sht.cell(row=row, column=col).value = label + "," + str(angle) + "," + arm["road"] + "," + str(x*excelMapWidth/800) + "," + str(y*excelMapHeight/800)    ### convert coords to fit a 500x500 map
                #sht.cell(row=row , column=col+1).value = angle
                #sht.cell(row=row, column=col + 2).value = arm["road"]
                print("outputting", label + "," + str(angle) + "," + arm["road"] + "," + str(x*excelMapWidth/800) + "," + str(y*excelMapHeight/800))
                col+=1
            row+=1
        folder = os.path.dirname(os.path.abspath(__file__))
        #folder = os.path.join(folder, "Runs\\")
        siteImg.save(folder + "/" + key + " with arm labels" + ".png")
        img2 = Image.open(key + " with arm labels.png")
        img2 = img2.resize((500, 500), Image.ANTIALIAS)
        excelMapImage = openpyxl.drawing.image.Image(img2)
        sht = wb.get_sheet_by_name("Maps")
        sht.add_image(excelMapImage, "B" + str(2 + count))

        count+=1

    count = 0
    ###
    ### remove any groups that have no sites attached
    ###
    groups = {k:v for k,v in groups.items() if len(v["siteList"]) != 0}
    for index,key in enumerate(sorted(groups.keys(), key=lambda k:0 if k == "ALL" else int(k.replace("Group ","").strip()))):
        print(index,key)
        if key != "ALL":
            groups["Group " + str(index)] = groups.pop(key)
    ###
    ### output groups
    ###

    for key,group in sorted(groups.items(), key=lambda x:0 if x[0] == "ALL" else int(x[0].replace("Group ","").strip())):
        print("key,value",key,group)
        download_group_map(key)
        img2 = Image.open(key + ".png")
        excelMapImage = openpyxl.drawing.image.Image(img2)
        sht = wb.get_sheet_by_name("Maps")
        sht.add_image(excelMapImage, "G" + str(2 + count))
        sht.cell(row=count + 2, column=6).value = key

        sht = wb.get_sheet_by_name("Data")
        sht.cell(row=count + 100, column= 1).value = key
        for i,site in enumerate(group["coords"]):
            armData = []
            siteName = site[0]
            siteLatLon = sites[siteName]["latlon"]
            for armLabel,arm in sorted(sites[siteName]["Arms"].items()):
                armLatLon = arm["latlon"]
                pixelDistance = mapmanager.pixelDistance(siteLatLon[0],siteLatLon[1],armLatLon[0],armLatLon[1],group["zoom"])
                armData.append(armLabel)
                armData.append(pixelDistance[0])
                armData.append((pixelDistance[1]))
            sht.cell(row=count + 100, column=i+2).value = ",".join(map(str,site)) + "," + ",".join(map(str,armData))

        ###
        ### create the overview map showing the sites in the group in red
        ### and the other sites in black
        ###

        img2 = Image.open("overview.png").convert('RGB')
        drawimage = ImageDraw.Draw(img2)
        fnt = ImageFont.truetype("arial", size=18)
        sitesInGroup = group["siteList"]
        if len(sitesInGroup) > 0:
            print("sitesingroup",sitesInGroup)
            siteList = [(site[0], site[1], site[2]) for site in get_all_site_details()]
            for site in siteList:
                if site[0] in sitesInGroup:
                    colour = "red"
                else:
                    colour  = "black"
                x, y = mapmanager.get_coords(overview_map_details[1], (site[1], site[2]), overview_map_details[2], size=1280)
                drawimage.ellipse([x - 15 - outline, y - 15 - outline, x + 15 + outline, y + 15 + outline], outline=colour,fill=colour)
                drawimage.ellipse([x - 15, y - 15, x + 15, y + 15], outline="white", fill="white")
                drawimage.text((x - 8, y - 9), text=site[0].split(" ")[1], font=fnt, fill="black")
            img2.save(key  + "_sites.png")
            img2 = Image.open(key  + "_sites.png")
            excelMapImage = openpyxl.drawing.image.Image(img2)
            sht = wb.get_sheet_by_name("Maps")
            sht.add_image(excelMapImage, "H" + str(2 + count))
            count += 1

    print("saving")
    try:

        wb.save(filename=file + ".xlsm")
    except PermissionError as e:
        messagebox.showinfo(message="file is already open, please close file and retry exporting")
        return

def get_all_site_coords():
    coords = []
    for _,site in project["sites"].items():
        if "J" in site["surveys"]:
            if not site["surveys"]["J"]["latlon"] in coords:
                coords.append(site["surveys"]["J"]["latlon"])
    return coords

def get_all_site_details():
    print("in get site details",sites)#
    print([site["Site Name"] + site["latlon"] for k, site in sites.items()])

    return [site["Site Name"] + site["latlon"] for k, site in sites.items()]

def download_all_individual_site_maps():
    for siteName,site in project["sites"].items():
        print("site is",site)
        for surveyType,survey in site["surveys"].items():
            print("survety is",survey)
            x,y = survey["latlon"]
            map = mapmanager.load_high_def_map_with_labels(x,y,survey["zoom"],imageType=survey["imageType"])
            survey["image"] = map.resize((800, 800), Image.ANTIALIAS)
            map.save(siteName + ".png")

def download_overview_map():
    global overview_map_details
    points = get_all_site_coords()
    print("in overview map points are",points)
    overview_map_details = mapmanager.load_overview_map_without_street_labels(points)
    overview_map_details[0].save("overview.png")
    #overview_map_details[0].show()
    print("centre of overview map is",overview_map_details[1],"zoom is",overview_map_details[2])

def download_group_map(groupName):
    grpList = groups[groupName]["siteList"]
    if len(grpList)== 0:
        return False
    groups[groupName]["coords"] = []
    print("grplist is",grpList)
    coordsList = [(site[1],site[2]) for site in get_all_site_details() if site[0] in grpList] ### coordsList is the list of lat/lons for each site in the group
    print("coordslist is",coordsList)
    map_details = mapmanager.load_overview_map_without_street_labels(coordsList)
    groups[groupName]["zoom"] = map_details[2]
    map_details[0].save(groupName + ".png")
    img2 = Image.open(groupName + ".png").convert('RGB')
    drawimage = ImageDraw.Draw(img2)
    fnt = ImageFont.truetype("arial", size=18)
    siteList = [(site[0],site[1],site[2]) for site in get_all_site_details() if site[0] in grpList]
    for site in siteList:
        x,y = mapmanager.get_coords(map_details[1],(site[1],site[2]),map_details[2],size=1280)
        groups[groupName]["coords"].append([site[0],x,y])
        drawimage.ellipse([x - 15- outline, y - 15- outline, x + 15+ outline, y + 15+ outline], outline="Black",fill = "black")
        drawimage.ellipse([x - 15, y - 15, x + 15, y + 15], outline="white",fill = "white")
        drawimage.text((x-8, y-9), text=site[0].split(" ")[1],font=fnt, fill="black")
    img2.save(groupName + ".png")

def road_orientation(angle):
    if angle > 360:
        angle-=360
    if angle>=348.75:
        return "N"
    elif angle < 11.25:
        return "N"
    elif angle>=11.25 and angle < 33.75:
        return "NNE"
    elif angle>=33.75 and angle < 56.25:
        return "NE"
    elif angle>=56.25 and angle < 78.75:
        return "ENE"
    elif angle>=78.75 and angle < 101.75:
        return "E"
    elif angle>=101.75 and angle < 123.75:
        return "ESE"
    elif angle>=123.75 and angle < 146.25:
        return "SE"
    elif angle>=146.25 and angle < 168.75:
        return "SSE"
    elif angle>=168.75 and angle < 191.25:
        return "S"
    elif angle>=191.25 and angle < 213.75:
        return "SSW"
    elif angle>=213.75 and angle < 236.25:
        return "SW"
    elif angle>=236.25 and angle < 258.75:
        return "WSW"
    elif angle>=258.75 and angle < 281.25:
        return "W"
    elif angle>=281.25 and angle < 303.75:
        return "WNW"
    elif angle>=303.75 and angle < 326.25:
        return "NW"
    elif angle>=326.25 and angle < 348.75:
        return "NNW"


#load_project()
#save_project_to_pickle("test.pkl")
#load_project_from_pickle("test.pkl")
#save_project_to_pickle("test.pkl")


#centre = (55.91009503466296, -3.501137500000034)


#export_to_excel()
#print(get_next_site("Site 1"))
#get_all_individual_site_maps()
#get_overview_map()

#register_arm_details("Site 1","A",320,320)

#import_site_details_from_excel()
#exit()
#
#load_project()
#groups = {}
#groups["testgroup2"] = ["Site 1","Site 2","Site 3"]
#download_group_map("testgroup2")

#import_site_details_from_excel()
#print(get_project_details())

#import_site_details_from_excel()

#import_site_details_from_excel()

#print(siteDetails[["Lat","Lon"]].values.tolist())